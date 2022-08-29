def email(exl_file):
    from openpyxl import load_workbook
    book = load_workbook(exl_file)
    sh = book.active
    
    class person:
        def __init__(self, name, surname, gender):
            self.name = name
            self.surname = surname
            self.gender = gender
    
    for i in range(2, sh.max_row+1):
        for j in range(1, sh.max_column+1):
            cell_obj = sh.cell(row=i, column=j)
            if j == 1 :
                first_name = cell_obj.value
            elif j == 2 :
                nick_name = cell_obj.value
            elif j == 3 :
                gnd = cell_obj.value
            elif j == 4 :
                data = person(first_name, nick_name, gnd)      
                if gnd == 'female':
                   cell_obj.value = ('\n' + 'Miss.' + data.name +'.'+ data.surname + '@email.com')
                elif gnd == 'male':
                    cell_obj.value = ('\n' + 'Mr.' + data.name +'.'+ data.surname + '@email.com')
    
        book.save(filename=exl_file)

try:
    email(r'.xlsx')  # <===  """"enter your excel file with the absolute path """"
except:
    print('FileNotFoundError')
